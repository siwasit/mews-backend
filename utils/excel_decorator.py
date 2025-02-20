from openpyxl import load_workbook
from openpyxl.styles import Border, Side

class ExcelDecorator:
    def __init__(self, filepath: str):
        self.filepath = filepath

    def decorate(self):
        wb = load_workbook(self.filepath)
        ws = wb.active
        ws.delete_rows(1)

        # Define border style (thin black lines)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Find the last row with content
        last_row = ws.max_row  # This will get the last row with any content

        # Apply borders from A8 to L (all rows with content)
        for row in range(7, last_row + 1):  # Start from row 7
            for col in range(1, 13):  # Columns A (1) to L (12)
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

        # Save the updated Excel file
        wb.save(self.filepath)
        wb.close()

class FirstRowDeleter:
    def __init__(self, filepath: str, sheet_name: str):
        self.filepath = filepath
        self.sheet_name = sheet_name

    def del_first_row(self):
        wb = load_workbook(self.filepath)    
        
        if self.sheet_name not in wb.sheetnames:
            print(f"Sheet {self.sheet_name} not found or not visible. Creating a temporary sheet.")
            # Create a temporary sheet if necessary
            wb.create_sheet(title=self.sheet_name)
            ws = wb[self.sheet_name]
            ws['A1'] = "Temporary sheet for deletion operation"
            wb.save(self.filepath)
            wb.close()
            return

        ws = wb[self.sheet_name]

        # Check if the sheet is not empty before deleting the first row
        if ws.max_row > 1:
            ws.delete_rows(1)

        wb.save(self.filepath)
        wb.close()

class Decorator:
    def __init__(self, filepath: str, sheet_name: str):
        self.filepath = filepath
        self.sheet_name = sheet_name

    def decorate(self):
        wb = load_workbook(self.filepath)
        ws = wb[self.sheet_name]  # Get the sheet by name

        # Define border style
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Find the last row with content
        last_row = ws.max_row + 1

        # Apply borders from A8 (row 8, as range starts from 7) to L (all rows with content)
        for row in range(7, last_row + 1):  # Correct row range
            for col in range(1, 13):  # Columns A to L
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

        wb.save(self.filepath)
        wb.close()